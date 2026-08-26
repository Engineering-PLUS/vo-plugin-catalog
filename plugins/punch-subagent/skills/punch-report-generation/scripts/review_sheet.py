#!/usr/bin/env python3
"""
review_sheet.py, the bulk-edit path for a rendered punch report.

The skill documents this tool but it was NOT shipped in the NVA06B pipeline
package, so bulk revision had no working path. This is that tool.

Why it exists: a rendered .docx is a compiled artifact. Small edits are fine in
Word (item headings use Word's own numbering, so deleting an item renumbers the
rest). But rewording, dropping, reordering or adding several items by hand in the
docx is error prone and gets overwritten on the next render. Instead, round-trip
through a spreadsheet.

    python3 review_sheet.py export build -o Report-Review.xlsx
    #   ... reviewer edits the yellow columns in Excel ...
    python3 review_sheet.py import build Report-Review.xlsx
    node scripts/gen_report.js build

One row per item. Yellow cells are reviewer-owned and editable, grey cells are
generated and are ignored on import, so photo paths and sheet clips cannot be
corrupted by editing the sheet. A timestamped .bak.json is written before any
change is applied.

Reviewer-owned columns:
    Include?          N drops the item from the report
    Order             spaced by 10, change to reorder
    Title
    Description
    Corrective Action
    Editor Note       cleared to remove the red internal box

PlanGrid ref is the permanent link back to the source and is never renumbered.
A new row with a blank PlanGrid ref inserts a new item (photo-less).
"""
import argparse
import datetime
import json
import os
import shutil
import sys

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl required:  pip install openpyxl --break-system-packages")

MASTER = "master_report_items.json"

EDITABLE = ["Include?", "Order", "Title", "Description", "Corrective Action", "Editor Note"]
GENERATED = ["PlanGrid ref", "Sheet", "Photos", "Origin", "Confidence", "Precedent basis",
             "Field engineer's original note"]
HEADERS = ["Include?", "Order", "PlanGrid ref", "Title", "Description", "Corrective Action",
           "Editor Note", "Sheet", "Photos", "Origin", "Confidence", "Precedent basis",
           "Field engineer's original note"]

YELLOW = PatternFill("solid", fgColor="FFF2CC")
GREY = PatternFill("solid", fgColor="EDEDED")
HEAD = PatternFill("solid", fgColor="44546A")


def export(build, out):
    master = json.load(open(os.path.join(build, MASTER)))
    wb = Workbook()
    ws = wb.active
    ws.title = "Punch items"

    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    for i, m in enumerate(master):
        ws.append([
            "Y",
            (i + 1) * 10,
            m.get("plangrid_ref", ""),
            m.get("title", ""),
            m.get("description", ""),
            m.get("corrective_action", ""),
            m.get("editor_note") or "",
            m.get("sheet_display", ""),
            len(m.get("photo_paths", [])),
            m.get("origin", ""),
            m.get("confidence", ""),
            m.get("precedent_note") or "",
            m.get("jim_original_text") or "",
        ])

    widths = {"A": 9, "B": 7, "C": 12, "D": 34, "E": 62, "F": 52, "G": 52,
              "H": 30, "I": 8, "J": 16, "K": 11, "L": 44, "M": 34}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 92
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Arial", size=9)
            cell.fill = YELLOW if HEADERS[c - 1] in EDITABLE else GREY

    # legend
    lr = ws.max_row + 2
    ws.cell(row=lr, column=1, value="Yellow = editable by reviewer.").font = Font(bold=True, name="Arial", size=9)
    ws.cell(row=lr + 1, column=1, value="Grey = generated, edits here are ignored on import.").font = Font(name="Arial", size=9)
    ws.cell(row=lr + 2, column=1, value="Include? = N drops the item. Order reorders. Blank PlanGrid ref on a new row inserts an item.").font = Font(name="Arial", size=9)

    wb.save(out)
    print(f"exported {len(master)} items -> {out}")


def do_import(build, xlsx):
    path = os.path.join(build, MASTER)
    master = json.load(open(path))
    by_ref = {m.get("plangrid_ref"): m for m in master}

    ws = load_workbook(xlsx).active
    head = [c.value for c in ws[1]]
    idx = {h: head.index(h) for h in HEADERS if h in head}

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or all(v in (None, "") for v in r):
            continue
        if str(r[0]).startswith(("Yellow", "Grey", "Include?")):
            continue
        rows.append(r)

    kept, dropped, inserted = [], 0, 0
    for r in rows:
        def g(col, default=""):
            i = idx.get(col)
            return default if i is None or i >= len(r) or r[i] is None else r[i]

        if str(g("Include?", "Y")).strip().upper().startswith("N"):
            dropped += 1
            continue

        ref = str(g("PlanGrid ref")).strip()
        base = by_ref.get(ref)
        if base is None:
            base = {"plangrid_ref": ref or "", "photo_paths": [], "photo_titles": [],
                    "sheet_display": g("Sheet"), "sheet_name": "", "sheet_description": "",
                    "location": "Not recorded in PlanGrid, see sheet reference",
                    "origin": "reviewer_added", "confidence": "n/a",
                    "precedent_note": None, "jim_original_text": None, "status": "open",
                    "photo_date": None}
            inserted += 1
        else:
            base = dict(base)

        # only reviewer-owned fields are read back
        base["title"] = str(g("Title")).strip()
        base["description"] = str(g("Description")).strip()
        base["corrective_action"] = str(g("Corrective Action")).strip()
        en = str(g("Editor Note")).strip()
        base["editor_note"] = en or None

        try:
            order = float(g("Order", 1e9))
        except (TypeError, ValueError):
            order = 1e9
        kept.append((order, base))

    kept.sort(key=lambda t: t[0])
    out = []
    for n, (_, m) in enumerate(kept, 1):
        m["display_number"] = n
        out.append(m)

    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    shutil.copy2(path, f"{path}.{stamp}.bak.json")
    json.dump(out, open(path, "w"), indent=2)

    print(f"backup   : {os.path.basename(path)}.{stamp}.bak.json")
    print(f"imported : {len(out)} items (dropped {dropped}, inserted {inserted})")
    print("re-render with:  node scripts/gen_report.js <build_dir>")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("mode", choices=["export", "import"])
    ap.add_argument("build")
    ap.add_argument("xlsx", nargs="?")
    ap.add_argument("-o", "--out", default="Report-Review.xlsx")
    a = ap.parse_args()
    if a.mode == "export":
        export(a.build, a.out)
    else:
        if not a.xlsx:
            sys.exit("import requires the .xlsx path")
        do_import(a.build, a.xlsx)


if __name__ == "__main__":
    main()
